import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WINE="5C0D16"; GOLD="B58A4E"; YELL="FFF2CC"; LGREY="F2ECE0"
white=Font(name="Arial",color="FFFFFF",bold=True,size=10)
hdrfill=PatternFill("solid",fgColor=WINE)
yfill=PatternFill("solid",fgColor=YELL)
zfill=PatternFill("solid",fgColor="FAF6EE")
titlefont=Font(name="Arial",bold=True,size=14,color=WINE)
subfont=Font(name="Arial",italic=True,size=9,color="7A1420")
cell=Font(name="Arial",size=10)
cellb=Font(name="Arial",size=10,bold=True)
thin=Side(style="thin",color="D9CDB8")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top")
ctr=Alignment(horizontal="center",vertical="center")

COLS=["Inimigo","Entrada","PA","Arma","Dano","PA arma","Alc.","Red. dano","Tática / observações",
      "PA rest.","Fadiga","Guarda","Cab","Tro","BD","BE","PD","PE","Status"]
WIDTHS=[17,20,4.5,24,17,6,9,17,46, 8,7,7, 4.5,4.5,4.5,4.5,4.5,4.5, 14]
TRACK_START=9  # índice (0-based) da 1ª coluna de acompanhamento (PA rest.)

def S(inimigo,entrada,pa,arma,dano,paa,alc,red,tat):
    return [inimigo,entrada,pa,arma,dano,paa,alc,red,tat]

CEN = [
 ("Cenário 1 — A Brecha na Muralha","Medieval · 15 exp · defesa da brecha",
  "Objetivo: segurar a brecha até o fim da rodada X / impedir que os saqueadores ultrapassem.",
  [S("Saqueador 1","Onda 1 (rod. 1)",10,"Faca / machadinha","1d4 / 1d8","2 / 4","1 m","—","Avança pela brecha; tenta ultrapassar os defensores."),
   S("Saqueador 2","Onda 1 (rod. 1)",10,"Faca / machadinha","1d4 / 1d8","2 / 4","1 m","—","Avança pela brecha; tenta ultrapassar os defensores."),
   S("Saqueador 3","Onda 1 (rod. 1)",10,"Faca / machadinha","1d4 / 1d8","2 / 4","1 m","—","Avança pela brecha; tenta ultrapassar os defensores."),
   S("Saqueador 4","Onda 2 (rod. 3)",10,"Faca / machadinha","1d4 / 1d8","2 / 4","1 m","—","Reforço da 2ª onda."),
   S("Bruto 1","Onda 2 (rod. 3)",8,"Maça pesada (contund.)","3d4","6","1 m","−1 tronco (peitoral leve)","Abre caminho na força."),
  ]),
 ("Cenário 2 — O Selo do Templo","Mukashi · 30 exp · furto/infiltração",
  "Objetivo: pegar o selo e sair. Guardas isolados = emboscada pelas costas; se o alarme sobe, eles se agrupam.",
  [S("Sentinela 1","Patrulha — Entrada",10,"Naginata curta (perf.)","1d6","4","1 m","—","Presa fácil por emboscada pelas costas."),
   S("Sentinela 2","Patrulha — Sala dos Sinos",10,"Naginata curta (perf.)","1d6","4","1 m","—","Presa fácil por emboscada pelas costas."),
   S("Sentinela 3","Patrulha — Jardim",10,"Naginata curta (perf.)","1d6","4","1 m","—","Presa fácil por emboscada pelas costas."),
   S("Capitão da Guarda","Santuário (junto ao selo)",10,"Katana (versátil)","1d8 CORT / 1d6 PERF","4","1 m","−1 (peitoral leve)","Fica ao lado do selo; lidera o reagrupamento se o alarme sobe."),
  ]),
 ("Cenário 3 — Correr no Ferro-Velho","Colapso · 50 exp · fuga/perseguição",
  "Objetivo do grupo: cruzar o portão. Regra: se ninguém está em corpo a corpo com um perseguidor no fim da rodada, o bando avança 4 m.",
  [S("Ferrasca 1","Largada (atrás do grupo)",10,"Cano de ferro (contund.)","3","4","1 m","—","Persegue em bloco; o mais rápido tenta agarrar quem ficou pra trás."),
   S("Ferrasca 2","Largada",10,"Cano de ferro (contund.)","3","4","1 m","—","Persegue em bloco."),
   S("Ferrasca 3","Largada",10,"Cano de ferro (contund.)","3","4","1 m","—","Persegue em bloco."),
   S("Ferrasca 4","Largada",10,"Cano de ferro (contund.)","3","4","1 m","—","Persegue em bloco."),
   S("Atirador Ferrasca","Retaguarda",10,"Besta leve","virote 1d4","4","7–10 m","—","Fica pra trás e atira em quem está na retaguarda."),
  ]),
 ("Cenário 4 — A Escolta da Chama","Medieval · 70 exp · escolta (proteger o Arauto)",
  "Objetivo: levar o Arauto vivo ao santuário. Inimigos em 3 ondas; o alvo prioritário deles é o Arauto (NPC).",
  [S("Bandido 1","Onda 1 (estrada)",10,"Espada curta (cort.)","1d8","4","1 m","—","Ataca antes da ponte."),
   S("Bandido 2","Onda 1 (estrada)",10,"Espada curta (cort.)","1d8","4","1 m","—","Ataca antes da ponte."),
   S("Bandido 3","Onda 1 (estrada)",10,"Espada curta (cort.)","1d8","4","1 m","—","Ataca antes da ponte."),
   S("Bandido 4","Onda 2 (ponte)",10,"Espada curta (cort.)","1d8","4","1 m","—","Gargalo da ponte."),
   S("Bandido 5","Onda 2 (ponte)",10,"Espada curta (cort.)","1d8","4","1 m","—","Gargalo da ponte."),
   S("Arqueiro","Onda 2 (ponte)",10,"Arco composto","flecha 1d4","4","7–10 m","—","Atira no Arauto de longe."),
   S("Chefe de bando","Onda 3 (santuário)",10,"Machado de batalha (cort.)","1d12","6","1–2 m","−2 (peitoral médio)","Tenta agarrar/derrubar o Arauto."),
  ]),
 ("Cenário 5 — O Coração do Colosso","Épico · 100 exp · combate de chefe",
  "Objetivo: destruir o Núcleo (20 de dano nele). REGIÕES DO COLOSSO valem 20 cada (o dobro): use Cab=Cabeça-cristal, Tro=NÚCLEO, BD/BE=braços, PD/PE=Base.",
  [S("Colosso de Pedra","Chefe (desde o início)",14,"Esmagada / Varredura / Tremor","3d4 / 2d4 / —","6 / 4 / —","2–3 m","−3 pele de pedra (mín 1); Núcleo exposto = sem redução","Núcleo blindado (ataques com desvantagem) até destruir 1 braço (20 dano). Esmagada acerta +1 espaço adjacente; Varredura empurra 3 m; Tremor 1×/combate (todos: cai + −2 PA). Move 4 m."),
   S("Fragmento 1","Surge na rodada 2",8,"Soco de pedra (improv. pesada)","4","4","1 m","—","10 PV por região. Distrai o grupo."),
   S("Fragmento 2","Surge na rodada 2",8,"Soco de pedra (improv. pesada)","4","4","1 m","—","10 PV por região. Distrai o grupo."),
   S("Fragmento 3","Surge na rodada 4",8,"Soco de pedra (improv. pesada)","4","4","1 m","—","10 PV por região. Distrai o grupo."),
   S("Fragmento 4","Surge na rodada 4",8,"Soco de pedra (improv. pesada)","4","4","1 m","—","10 PV por região. Distrai o grupo."),
  ]),
 ("Cenário 6 — O Vale dos Atiradores","Colapso · 50 exp · incursão sob fogo (TESTE DE COBERTURA)",
  "Objetivo do grupo: alcançar o relé no topo. Atiradores no alto (elevação = atiram com vantagem) e em cobertura; Cães emboscam no meio.",
  [S("Olho de Chumbo 1","Espinhaço (elevação +2m)",10,"Rifle","bala 1d8","6","3–20 m","cobertura própria","Atira com vantagem (elevação) de cobertura completa; abaixa entre tiros. Silenciar = flanquear a rampa ou pegar a janela em que se expõe."),
   S("Olho de Chumbo 2","Espinhaço (elevação +2m)",10,"Rifle","bala 1d8","6","3–20 m","cobertura própria","Idem — outro espinhaço."),
   S("Olho de Chumbo 3","Espinhaço (elevação +2m)",10,"Rifle","bala 1d8","6","3–20 m","cobertura própria","Idem."),
   S("Cão da Gorja 1","Emboscada (manilha, meio)",10,"Cano de ferro (contund.)","3","4","1 m","—","Surge no terreno difícil; agarra e prende isolados."),
   S("Cão da Gorja 2","Emboscada (manilha, meio)",10,"Cano de ferro (contund.)","3","4","1 m","—","Idem."),
   S("Cão da Gorja 3","Emboscada (manilha, meio)",10,"Cano de ferro (contund.)","3","4","1 m","—","Idem."),
   S("Cão da Gorja 4","Emboscada (manilha, meio)",10,"Cano de ferro (contund.)","3","4","1 m","—","Idem."),
   S("Vigia-mor (opcional)","Retaguarda",12,"Besta pesada","virote 1d8","6","8–12 m","—","Coordena os atiradores; some se um espinhaço cair."),
  ]),
]

wb=openpyxl.Workbook()
# --- Instruções ---
ins=wb.active; ins.title="Instruções"
ins.sheet_view.showGridLines=False
ins["A1"]="Marca de Sangue — Controle de Inimigos (playtest)"; ins["A1"].font=Font(name="Arial",bold=True,size=16,color=WINE)
linhas=[
 "",
 "Uma aba por cenário. Cada linha = um inimigo individual. As colunas claras já vêm preenchidas (referência);",
 "as colunas AMARELAS são para você preencher durante o jogo.",
 "",
 "Colunas de referência: Inimigo · Entrada (onda/rodada/local) · PA (base do turno) · Arma · Dano · PA da arma ·",
 "Alcance ideal · Red. dano (armadura) · Tática/observações.",
 "",
 "Colunas de acompanhamento (amarelas):",
 "   • PA rest. — pontos de ação restantes no turno.",
 "   • Fadiga — fadiga acumulada. Penalidade: −1 PA a cada 5 a partir de 10; 50 = inconsciente.",
 "   • Guarda — nº de guardas levantadas (defesa/esquiva com vantagem; 2 PA fora de engajamento, 3 engajado).",
 "   • Cab / Tro / BD / BE / PD / PE — dano acumulado em cada membro (cabeça, tronco, braços, pernas).",
 "   • Status — ativo · atordoado · caído · incapacitado · fugiu · morto.",
 "",
 "Vida (PV) — MODELO NOVO (playtest 1): cada membro tem 10 espaços. Enche de dano → vira PROFUNDO",
 "(incapacitado; só cura com tratamento). Dano além disso → PERMANENTE. Todo dano recebido gera 1 fadiga/ponto.",
 "HP de MINION (para o mestre): fraco = cai com 10 de dano total · médio = 20 · forte = 30.",
 "EXCEÇÃO — Colosso (Cenário 5): cada região aguenta 20 (o dobro). Use Cab=Cabeça-cristal, Tro=NÚCLEO,",
 "BD/BE=braços, PD/PE=Base. Vencer = 20 de dano no Núcleo (que só abre ao destruir um braço).",
 "",
 "Dica: dá para imprimir (paisagem) ou usar na tela. O cabeçalho e a coluna do nome ficam congelados ao rolar.",
]
for i,t in enumerate(linhas,start=2):
    ins[f"A{i}"]=t; ins[f"A{i}"].font=Font(name="Arial",size=10, bold=(t.endswith(":")))
ins.column_dimensions["A"].width=115

def make_tab(nome_aba, titulo, sub, obj, rows):
    ws=wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines=False
    ncol=len(COLS)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncol)
    ws.cell(1,1,titulo).font=titlefont
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncol)
    ws.cell(2,1,sub).font=subfont
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=ncol)
    oc=ws.cell(3,1,obj); oc.font=Font(name="Arial",size=9,color="333333"); oc.alignment=wrap
    ws.row_dimensions[3].height=28
    # header
    hr=5
    for j,h in enumerate(COLS,start=1):
        c=ws.cell(hr,j,h); c.font=white; c.fill=hdrfill; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
    for j,w in enumerate(WIDTHS,start=1):
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[hr].height=26
    # rows
    r=hr+1
    for row in rows:
        for j in range(ncol):
            c=ws.cell(r,j+1)
            c.border=border; c.font=cell
            if j < len(row):
                c.value=row[j]
            if j in (2,5,6,7): c.alignment=ctr  # PA, PA arma, Alc, Red
            elif j==0: c.alignment=Alignment(vertical="center"); c.font=cellb
            elif j==8: c.alignment=wrap
            else: c.alignment=ctr
            if j>=TRACK_START:
                c.fill=yfill
            elif r%2==0:
                c.fill=zfill
        ws.row_dimensions[r].height=34
        r+=1
    ws.freeze_panes="B6"
    ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    return ws

for (titulo,sub,obj,rows) in CEN:
    aba = titulo.split("—")[0].strip().replace("Cenário","Cen.")  # "Cen. 1"
    make_tab(aba, titulo, sub, obj, rows)

out="/sessions/dreamy-amazing-ritchie/mnt/outputs/Marca-de-Sangue-Controle-de-Inimigos.xlsx"
wb.save(out)
print("salvo:", out, "abas:", wb.sheetnames)
